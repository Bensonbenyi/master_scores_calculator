from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
from datetime import datetime

app = Flask(__name__, template_folder='html_files')

# 配置数据库
import os
database_url = os.environ.get('DATABASE_URL', 'sqlite:///users.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')  

# 配置上传文件
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'txt', 'doc', 'docx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 创建上传目录（如果不存在）
os.makedirs(os.path.join(app.root_path, UPLOAD_FOLDER), exist_ok=True)

# 创建数据库实例
db = SQLAlchemy(app)

# 创建迁移实例
migrate = Migrate(app, db)

# 用户模型
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), default='student')  # 'student' 或 'teacher'
    full_name = db.Column(db.String(100))
    student_id = db.Column(db.String(20))
    major = db.Column(db.String(100))
    
    # 成绩相关字段
    academic_score = db.Column(db.Float, default=0)
    academic_talent_score = db.Column(db.Float, default=0)
    comprehensive_score = db.Column(db.Float, default=0)
    final_score = db.Column(db.Float, default=0)
    
    # 学术专长细分项成绩
    paper_score = db.Column(db.Float, default=0)
    patent_score = db.Column(db.Float, default=0)
    competition_national_score = db.Column(db.Float, default=0)
    competition_provincial_score = db.Column(db.Float, default=0)
    csp_score = db.Column(db.Float, default=0)
    innovation_project_score = db.Column(db.Float, default=0)
    
    # 综合表现细分项成绩
    honor_score = db.Column(db.Float, default=0)
    social_work_score = db.Column(db.Float, default=0)
    volunteer_score = db.Column(db.Float, default=0)
    volunteer_hours = db.Column(db.Integer, default=0)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

# 不再使用 db.create_all()，改用 Flask-Migrate 管理数据库结构
# 使用命令：flask db init, flask db migrate, flask db upgrade

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['username'] = username
            session['role'] = user.role
            
            if user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('transition_page'))
        else:
            return '用户名或密码错误'
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    role = request.args.get('role', 'student')
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role', 'student')
        
        # 检查用户名是否已存在
        if User.query.filter_by(username=username).first():
            return '用户名已存在'
        
        # 如果是教师注册，需要验证教师码
        if role == 'teacher':
            teacher_code = request.form.get('teacher_code')
            # 这里可以设置一个固定的教师验证码，实际应用中可能需要更复杂的验证方式
            if teacher_code != 'teacher123':
                return '教师验证码错误'
        
        # 创建新用户
        user = User(username=username, role=role)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        return redirect(url_for('login'))
    
    return render_template('register.html', role=role)

@app.route('/transition')
def transition_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # 检查用户角色，教师应该重定向到教师面板
    if session.get('role') == 'teacher':
        return redirect(url_for('teacher_dashboard'))
    
    username = session['username']
    return render_template('transition.html', username=username)

@app.route('/profile', methods=['GET', 'POST'])
def profile_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    user = User.query.filter_by(username=username).first()
    
    return render_template('profile.html', username=username, profile=user)

@app.route('/save_profile', methods=['POST'])
def save_profile():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    user = User.query.filter_by(username=username).first()
    
    if user:
        user.full_name = request.form.get('full_name')
        user.student_id = request.form.get('student_id')
        user.major = request.form.get('major')
        
        db.session.commit()
        flash('个人信息保存成功！', 'success')
    
    return redirect(url_for('profile_page'))

@app.route('/student')
def student_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    return render_template('stu_page.html', username=username)

@app.route('/calculate_score', methods=['POST'])
def calculate_score():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    
    try:
        # 获取学业成绩，添加错误处理
        academic_score_str = request.form.get('academic_score', '0')
        if not academic_score_str or not academic_score_str.replace('.', '').isdigit():
            flash('请输入有效的学业成绩', 'error')
            return redirect(url_for('student_page'))
        
        academic_score = float(academic_score_str)
        if academic_score < 0 or academic_score > 100:
            flash('学业成绩应在0-100之间', 'error')
            return redirect(url_for('student_page'))
    except ValueError:
        flash('学业成绩格式错误', 'error')
        return redirect(url_for('student_page'))
    
    # 处理文件上传
    uploaded_files = []
    if 'proof_files' in request.files:
        files = request.files.getlist('proof_files')
        for file in files:
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # 使用用户名作为子文件夹，防止文件名冲突
                user_folder = os.path.join(app.config['UPLOAD_FOLDER'], username)
                os.makedirs(os.path.join(app.root_path, user_folder), exist_ok=True)
                file_path = os.path.join(user_folder, filename)
                file.save(os.path.join(app.root_path, file_path))
                uploaded_files.append(file_path)
    
    # 计算学术专长成绩
    academic_talent_score = 0
    
    # 学术论文加分
    paper_score = 0
    for paper in request.form.getlist('academic_paper'):
        if paper == 'nature_science_first':
            paper_score += 12
        elif paper == 'ccf_a_first':
            paper_score += 1.5
        elif paper == 'ccf_a_second':
            paper_score += 0.75
        elif paper == 'ccf_a_third':
            paper_score += 0.45
        elif paper == 'ccf_a_fourth':
            paper_score += 0.3
        elif paper == 'ccf_c_first':
            paper_score += 0.25
        elif paper == 'ccf_c_second':
            paper_score += 0.125
        elif paper == 'ccf_c_third':
            paper_score += 0.075
    academic_talent_score += paper_score
    session['paper_score'] = paper_score
    
    # 发明专利加分
    patent_score = 0
    for patent in request.form.getlist('patent'):
        if patent == 'patent_first_2':
            patent_score += 1.8
        elif patent == 'patent_first_4':
            patent_score += 1.5
        elif patent == 'patent_third':
            patent_score += 0.45
        elif patent == 'patent_fourth':
            patent_score += 0.3
    academic_talent_score += patent_score
    session['patent_score'] = patent_score
    
    # 国家级竞赛加分
    competition_national_score = 0
    for competition in request.form.getlist('competition_national'):
        if competition == 'a_plus_first_team':
            competition_national_score += 4
        elif competition == 'a_first_team':
            competition_national_score += 3.375
        elif competition == 'a_minus_first_team':
            competition_national_score += 2.6666
        elif competition == 'a_second_team':
            competition_national_score += 2.25
        elif competition == 'a_minus_second_team':
            competition_national_score += 1.5
        elif competition == 'a_third_team':
            competition_national_score += 1.35
        elif competition == 'a_minus_third_team':
            competition_national_score += 0.9
        elif competition == 'a_second_individual':
            competition_national_score += 2.25
        elif competition == 'a_third_individual':
            competition_national_score += 1.35
    academic_talent_score += competition_national_score
    session['competition_national_score'] = competition_national_score
    
    # 省级竞赛加分
    competition_provincial_score = 0
    for competition in request.form.getlist('competition_provincial'):
        if competition == 'a_first_team':
            competition_provincial_score += 1.2
        elif competition == 'a_minus_first_team':
            competition_provincial_score += 0.8
        elif competition == 'a_second_team':
            competition_provincial_score += 0.6
        elif competition == 'a_minus_second_team':
            competition_provincial_score += 0.4
        elif competition == 'a_third_team':
            competition_provincial_score += 0.27
        elif competition == 'a_first_individual':
            competition_provincial_score += 1.2
        elif competition == 'a_second_individual':
            competition_provincial_score += 0.6
    academic_talent_score += competition_provincial_score
    session['competition_provincial_score'] = competition_provincial_score
    
    # CCF CSP认证加分
    csp_score = 0
    for csp in request.form.getlist('ccf_csp'):
        if csp == 'csp_400':
            csp_score += 4
        elif csp == 'csp_320_399':
            csp_score += 3
        elif csp == 'csp_280_319':
            csp_score += 2
    academic_talent_score += csp_score
    session['csp_score'] = csp_score
    
    # 创新创业训练项目加分
    innovation_project_score = 0
    for project in request.form.getlist('innovation_project'):
        if project == 'national_leader':
            innovation_project_score += 1
        elif project == 'national_member':
            innovation_project_score += 0.3
        elif project == 'provincial_leader':
            innovation_project_score += 0.5
        elif project == 'provincial_member':
            innovation_project_score += 0.2
        elif project == 'school_leader':
            innovation_project_score += 0.1
        elif project == 'school_member':
            innovation_project_score += 0.05
    academic_talent_score += innovation_project_score
    session['innovation_project_score'] = innovation_project_score
    
    # 限制学术专长成绩最高为12分
    academic_talent_score = min(academic_talent_score, 12)
    
    # 计算综合表现加分
    comprehensive_score = 0
    
    # 荣誉称号加分
    honor_score = 0
    for honor in request.form.getlist('honor'):
        if honor == 'good_student':
            honor_score = max(honor_score, 0.2)  # 同年度不累计
        elif honor == 'excellent_student':
            honor_score = max(honor_score, 0.2)  # 同年度不累计
        elif honor == 'excellent_cadre':
            honor_score = max(honor_score, 0.2)  # 同年度不累计
        elif honor == 'excellent_league_member':
            honor_score = max(honor_score, 0.2)  # 同年度不累计
        elif honor == 'red_flag_branch':
            honor_score += 0.1  # 集体荣誉可以加分
    
    comprehensive_score += honor_score
    session['honor_score'] = honor_score
    
    # 社会工作加分
    social_work_score = 0
    for work in request.form.getlist('social_work'):
        if work == 'class_monitor':
            social_work_score += 1
        elif work == 'deputy_monitor':
            social_work_score += 0.5
        elif work == 'study_committee':
            social_work_score += 0.485
        elif work == 'other_committee':
            social_work_score += 0.475
        elif work == 'student_union_chair':
            social_work_score += 1.5
        elif work == 'student_union_director':
            social_work_score += 1.0
        elif work == 'student_union_deputy':
            social_work_score += 0.7125
        elif work == 'student_union_member':
            social_work_score += 0.5
        elif work == 'club_president':
            social_work_score += 0.75
        elif work == 'club_vice_president':
            social_work_score += 0.5
        elif work == 'club_member':
            social_work_score += 0.5
    
    # 社会工作加分累计不超过2分
    social_work_score = min(social_work_score, 2)
    comprehensive_score += social_work_score
    session['social_work_score'] = social_work_score
    
    # 志愿服务加分
    volunteer_hours = int(request.form.get('volunteer_hours', 0))
    volunteer_score = 0
    
    if volunteer_hours >= 200:
        volunteer_score = 1
        # 每增加2小时加0.05分，最高不超过1分
        additional_hours = volunteer_hours - 200
        additional_score = min((additional_hours // 2) * 0.05, 1)
        volunteer_score += additional_score
    
    comprehensive_score += volunteer_score
    session['volunteer_score'] = volunteer_score
    
    # 限制综合表现加分最高为8分
    comprehensive_score = min(comprehensive_score, 8)
    
    # 计算最终成绩
    academic_weighted = academic_score * 0.8
    academic_talent_weighted = academic_talent_score  # 学术专长成绩，已限制最高为12分
    comprehensive_weighted = comprehensive_score  # 综合表现加分，已限制最高为8分
    
    final_score = academic_weighted + academic_talent_weighted + comprehensive_weighted
    
    # 保存成绩到用户模型
    user = User.query.filter_by(username=username).first()
    if user:
        user.academic_score = round(academic_weighted, 3)
        user.academic_talent_score = round(academic_talent_weighted, 3)
        user.comprehensive_score = round(comprehensive_weighted, 3)
        user.final_score = round(final_score, 3)
        
        # 保存学术专长细分项成绩
        user.paper_score = round(paper_score, 3)
        user.patent_score = round(patent_score, 3)
        user.competition_national_score = round(competition_national_score, 3)
        user.competition_provincial_score = round(competition_provincial_score, 3)
        user.csp_score = round(csp_score, 3)
        user.innovation_project_score = round(innovation_project_score, 3)
        
        # 保存综合表现细分项成绩
        user.honor_score = round(honor_score, 3)
        user.social_work_score = round(social_work_score, 3)
        user.volunteer_score = round(volunteer_score, 3)
        user.volunteer_hours = volunteer_hours
        
        db.session.commit()
    
    # 保存上传的文件路径到会话中，便于在信息页面显示
    session['uploaded_files'] = uploaded_files
    
    return render_template('stu_page.html', 
                          username=username,
                          final_score=round(final_score, 3),
                          academic_weighted=round(academic_weighted, 3),
                          academic_talent_weighted=round(academic_talent_weighted, 3),
                          comprehensive_weighted=round(comprehensive_weighted, 3))

@app.route('/student_info')
def student_info():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    user = User.query.filter_by(username=username).first()
    
    if not user:
        return redirect(url_for('login'))
    
    # 获取上传的文件列表
    uploaded_files = []
    user_folder = os.path.join(app.config['UPLOAD_FOLDER'], username)
    user_folder_path = os.path.join(app.root_path, user_folder)
    
    if os.path.exists(user_folder_path):
        for file in os.listdir(user_folder_path):
            if os.path.isfile(os.path.join(user_folder_path, file)) and allowed_file(file):
                # 构建相对于static目录的路径
                relative_path = os.path.join('uploads', username, file)
                uploaded_files.append(relative_path)
    
    # 学术专长成绩详情（从数据库获取，而不是session）
    academic_talent_details = [
        {"name": "学术论文", "score": round(user.paper_score or 0, 3)},
        {"name": "发明专利", "score": round(user.patent_score or 0, 3)},
        {"name": "国家级竞赛", "score": round(user.competition_national_score or 0, 3)},
        {"name": "省级竞赛", "score": round(user.competition_provincial_score or 0, 3)},
        {"name": "CCF CSP认证", "score": round(user.csp_score or 0, 3)},
        {"name": "创新创业训练项目", "score": round(user.innovation_project_score or 0, 3)}
    ]
    
    # 综合表现加分详情（从数据库获取，而不是session）
    comprehensive_details = [
        {"name": "荣誉称号", "score": round(user.honor_score or 0, 3)},
        {"name": "社会工作", "score": round(user.social_work_score or 0, 3)},
        {"name": "志愿服务", "score": round(user.volunteer_score or 0, 3)}
    ]
    
    # 确保值不为None
    final_score = user.final_score if user.final_score is not None else 0
    academic_weighted = user.academic_score if user.academic_score is not None else 0
    academic_talent_weighted = user.academic_talent_score if user.academic_talent_score is not None else 0
    comprehensive_weighted = user.comprehensive_score if user.comprehensive_score is not None else 0
    
    return render_template('stu_info.html',
                          username=username,
                          user=user,
                          final_score=final_score,
                          academic_weighted=academic_weighted,
                          academic_talent_weighted=academic_talent_weighted,
                          comprehensive_weighted=comprehensive_weighted,
                          academic_talent_details=academic_talent_details,
                          comprehensive_details=comprehensive_details,
                          uploaded_files=uploaded_files)

@app.route('/teacher/dashboard')
def teacher_dashboard():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('login'))
    
    username = session['username']
    
    # 获取排序参数
    sort_by = request.args.get('sort_by', 'final_score')
    order = request.args.get('order', 'desc')
    
    # 验证排序字段是否有效
    valid_sort_fields = ['final_score', 'academic_score', 'academic_talent_score', 'comprehensive_score']
    if sort_by not in valid_sort_fields:
        sort_by = 'final_score'
    
    # 构建排序表达式
    sort_expr = getattr(User, sort_by)
    if order == 'desc':
        sort_expr = sort_expr.desc()
    else:
        sort_expr = sort_expr.asc()
    
    # 获取学生列表（只查询学生角色的用户）
    students = User.query.filter_by(role='student').order_by(sort_expr).all()
    
    return render_template('teacher_dashboard.html', 
                          username=username, 
                          students=students, 
                          sort_by=sort_by, 
                          order=order)

@app.route('/teacher/student/<int:student_id>')
def student_detail(student_id):
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('login'))
    
    username = session['username']
    
    # 获取学生信息
    student = User.query.get_or_404(student_id)
    
    # 确保查看的是学生用户
    if student.role != 'student':
        return redirect(url_for('teacher_dashboard'))
    
    # 获取学生上传的文件列表
    uploaded_files = []
    user_folder = os.path.join(app.config['UPLOAD_FOLDER'], student.username)
    user_folder_path = os.path.join(app.root_path, user_folder)
    
    if os.path.exists(user_folder_path):
        for file in os.listdir(user_folder_path):
            if os.path.isfile(os.path.join(user_folder_path, file)) and allowed_file(file):
                # 构建相对于static目录的路径
                relative_path = os.path.join('uploads', student.username, file)
                uploaded_files.append(relative_path)
    
    # 学术专长成绩详情
    academic_talent_details = [
        {"name": "学术论文", "score": student.paper_score or 0},
        {"name": "发明专利", "score": student.patent_score or 0},
        {"name": "国家级竞赛", "score": student.competition_national_score or 0},
        {"name": "省级竞赛", "score": student.competition_provincial_score or 0},
        {"name": "CCF CSP认证", "score": student.csp_score or 0},
        {"name": "创新创业训练项目", "score": student.innovation_project_score or 0}
    ]
    
    # 综合表现加分详情
    comprehensive_details = [
        {"name": "荣誉称号", "score": student.honor_score or 0},
        {"name": "社会工作", "score": student.social_work_score or 0},
        {"name": "志愿服务", "score": student.volunteer_score or 0}
    ]
    
    return render_template('student_detail.html',
                          username=username,
                          student=student,
                          final_score=student.final_score,
                          academic_weighted=student.academic_score,
                          academic_talent_weighted=student.academic_talent_score,
                          comprehensive_weighted=student.comprehensive_score,
                          academic_talent_details=academic_talent_details,
                          comprehensive_details=comprehensive_details,
                          uploaded_files=uploaded_files)

@app.route('/teacher/export_excel')
def export_excel():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('login'))
    
    # 获取所有学生数据，按综合成绩降序排列
    students = User.query.filter_by(role='student').order_by(User.final_score.desc()).all()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生推免成绩排名"
    
    # 设置标题样式
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # 设置表头
    headers = [
        '排名', '姓名', '学号', '专业', '用户名',
        '综合成绩', '学业成绩', '学术专长成绩', '综合表现成绩',
        '学术论文', '发明专利', '国家级竞赛', '省级竞赛', 'CCF CSP认证', '创新创业训练项目',
        '荣誉称号', '社会工作', '志愿服务', '志愿服务时长(小时)'
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
    
    # 写入学生数据
    for row, student in enumerate(students, 2):
        rank = row - 1
        data = [
            rank,
            student.full_name or '未填写',
            student.student_id or '未填写',
            student.major or '未填写',
            student.username,
            student.final_score or 0,
            student.academic_score or 0,
            student.academic_talent_score or 0,
            student.comprehensive_score or 0,
            student.paper_score or 0,
            student.patent_score or 0,
            student.competition_national_score or 0,
            student.competition_provincial_score or 0,
            student.csp_score or 0,
            student.innovation_project_score or 0,
            student.honor_score or 0,
            student.social_work_score or 0,
            student.volunteer_score or 0,
            student.volunteer_hours or 0
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = center_alignment
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
    
    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'学生推免成绩排名_{timestamp}.xlsx'
    
    return send_file(temp_file.name, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/test')
def test_route():
    """测试路由"""
    return "应用正常运行！当前时间: " + str(datetime.now())

@app.route('/admin/view_data')
def admin_view_data():
    """管理员查看所有数据"""
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('login'))
    
    try:
        # 获取所有用户数据
        all_users = User.query.all()
        students = User.query.filter_by(role='student').all()
        teachers = User.query.filter_by(role='teacher').all()
        
        return render_template('admin_data.html', 
                             all_users=all_users,
                             students=students, 
                             teachers=teachers)
    except Exception as e:
        return f"数据库查询错误: {str(e)}", 500

@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('role', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, port=port, host='0.0.0.0')